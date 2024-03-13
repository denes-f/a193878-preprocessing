"""
Generates a list of scenarios from the HARA sheet.
"""
import copy
import math
import os
import openpyxl
import openpyxl.styles
from ConfigReader import ConfigReader


def main(mode):
    """
Generates a list of scenarios from the HARA sheet.
    """

    config_path = 'config.ini'
    config_reader = ConfigReader(config_path)

    hara_path = config_reader.get_value('Hara_Sheet', 'path')
    hara_sheet_name = config_reader.get_value('Hara_Sheet', 'sheet_name')
    hara_reader = HaraReader(hara_path, hara_sheet_name, config_reader, config_reader.get_int('Hara_Sheet', 'header_size'))

    scenario_template_path = config_reader.get_value('Scenario_Template', 'path')
    scenario_template_sheet_name = config_reader.get_value('Scenario_Template', 'sheet_name')

    scenario_writer = ScenarioWriter(config_reader, scenario_template_path, scenario_template_sheet_name, config_reader.get_int('Scenario_Template', 'header_size'), mode)

    while True:  # Iterating through the items in the HARA
        hara_item = hara_reader.get_next_item()
        if hara_item.item_id is None:
            break
        if not hara_item.relevant:
            continue
        scenario = Scenario(config_reader, hara_item)  # Converting HARA items to the Scenario list (using the config settings)
        scenario_writer.write(hara_item, scenario)  # Writing to the Scenario list
    scenario_writer.save()


class SteeringFault:
    """
Steering malfunction
    """

    def __init__(self, additional_unintended_angle, slew_rate):
        self.additional_unintended_angle = additional_unintended_angle
        self.slew_rate = slew_rate


class RearSteeringFault:
    """
Rear axle steering malfunction
    """

    def __init__(self, additional_unintended_angle, slew_rate):
        self.additional_unintended_angle = additional_unintended_angle
        self.slew_rate = slew_rate


class TorqueFault:
    """
E-motor torque malfunction
    """

    def __init__(self, torque_error_front=None, torque_error_rear=None, slew_rate=None):
        self.torque_error_front = torque_error_front
        self.torque_error_rear = torque_error_rear
        self.slew_rate = slew_rate

    def get_overall_torque(self):
        """
Calculates the overall torque
        :return: Returns the overall torque from the front and rear e-motors
        """
        torque_front = self.torque_error_front if isinstance(self.torque_error_front, float) else 0.0
        torque_rear = self.torque_error_rear if isinstance(self.torque_error_rear, float) else 0.0
        overall_torque = torque_front + torque_rear
        return overall_torque

    def losing_stability(self) -> bool:
        """
Checks the possibility of losing stability
        :return: Returns true if there is a possibility for losing stability
        """
        if self.torque_error_rear is not None:
            if self.torque_error_rear < 50:
                return True
        if self.torque_error_front is not None and self.torque_error_rear is not None:
            if self.torque_error_front < 0 and self.torque_error_rear > 0:
                return True
            if self.torque_error_front > 0 and self.torque_error_rear < 0:
                return True
        return False


class RideHeightFault:
    """
Ride height malfunction
    """

    def __init__(self, front_left=None, front_right=None, rear_left=None, rear_right=None, slew_rate=None):
        self.front_left = front_left
        self.front_right = front_right
        self.rear_left = rear_left
        self.rear_right = rear_right
        self.slew_rate = slew_rate

    @classmethod
    def all_wheels(cls, unintended_height, slew_rate):
        """
Unintended ride height malfunction for all wheels
        :param slew_rate: The rate in m/s at which the ride height fault is injected
        :param unintended_height: The unintended ride height adjustment in meters
        :return: Returns a RideHeightFault with all wheels having the same unintended ride height
        """
        return cls(unintended_height, unintended_height, unintended_height, unintended_height, slew_rate)


class BrakingFault:
    """
Hydraulic brake malfunction
    """

    def __init__(self, unintended_braking_torque):
        self.unintended_braking_torque = unintended_braking_torque


class ParkBrakeFault:
    """
Park brake malfunction
    """

    def __init__(self, park_brake_status):
        self.park_brake_status = park_brake_status
        raise NotImplementedError('Park brake faults are not implemented yet')


class VerySlowSteeringReaction:
    """
Applied steering reaction in degrees per second
    """

    def __init__(self, steering_rate_limit):
        self.steering_rate_limit = steering_rate_limit


class SlowSteeringReaction:
    """
Applied steering reaction in degrees per second
    """

    def __init__(self, steering_rate_limit):
        self.steering_rate_limit = steering_rate_limit


class BrakingReaction:
    """
Applied braking force in percentage
    """

    def __init__(self, braking):
        self.braking = braking

class FaultTolerantTime:
    """
Applied braking force in percentage
    """

    def __init__(self, ftti):
        self.ftti = ftti

class HaraItem:
    """
Type containing the properties of one line in the HARA
    """
    def __init__(self, item, location, slope, route, road_condition, engaged_gear, vehicle_speed, brake_pedal, hazard, relevant, comment):
        self.item_id = item
        self.location = location
        self.slope = slope
        self.route = route
        self.road_condition = road_condition
        self.engaged_gear = engaged_gear
        self.vehicle_speed = vehicle_speed
        self.brake_pedal = brake_pedal
        self.hazard = hazard
        self.relevant = relevant
        self.comment = comment


class Scenario:
    """
Converts a HARA item to a Scenario (using the config settings)
    """

    def __init__(self, config_reader, hara_item):
        self.config_reader = config_reader
        slope = hara_item.slope.lower()
        route = hara_item.route.lower()
        road_condition = hara_item.road_condition.lower()
        engaged_gear = hara_item.engaged_gear.lower()
        vehicle_speed = hara_item.vehicle_speed.lower()
        brake_pedal = hara_item.brake_pedal.lower()
        hazard = hara_item.hazard

        if slope == '-' or 'flat' in slope:
            road_gradient = config_reader.get_value('Slope', 'flat')
        elif 'slight' in slope:
            road_gradient = config_reader.get_value('Slope', 'slight_slope')
        elif 'down' in slope:
            road_gradient = config_reader.get_value('Slope', 'downhill')
        elif 'up' in slope:
            road_gradient = config_reader.get_value('Slope', 'uphill')
        else:
            raise Exception('Slope "{0}" not recognized in item {1}'.format(slope, hara_item.item_id))
        try:
            self.road_gradient = float(road_gradient)
        except ValueError:
            raise ValueError("Invalid road gradient '{0}' in config file, in Slope section".format(road_gradient))

        if vehicle_speed == '-' or 'stand' in vehicle_speed:
            speed = config_reader.get_value('Speed', 'standstill')
        elif 'very low' in vehicle_speed:
            speed = config_reader.get_value('Speed', 'very_low')
        elif 'low' in vehicle_speed:
            speed = config_reader.get_value('Speed', 'low')
        elif 'medium' in vehicle_speed:
            speed = config_reader.get_value('Speed', 'medium')
        elif 'high' in vehicle_speed:
            speed = config_reader.get_value('Speed', 'high')
        else:
            raise Exception('Speed "{0}" not recognized in item {1}'.format(vehicle_speed, hara_item.item_id))
        speed_list = speed.strip('[').strip(']').split(',')
        self.vehicle_speed = [.0] * len(speed_list)
        for i in range(len(speed_list)):
            try:
                # Currently reverse driving is not supported by the VSM model. Therefore positive speeds are used in all scenarios.
                self.vehicle_speed[i] = float(speed_list[i]) if engaged_gear != 'r' else 1 * float(speed_list[i])
            except ValueError:
                raise ValueError("Invalid speed thresholds in config file, '{0}' in Speed section".format(speed))

        if route == '-' or 'straight' in route or 'any' in route:
            self.road_radius = 'straight'
        elif 'curve' in route:
            if vehicle_speed == '-' or 'stand' in vehicle_speed or 'low' in vehicle_speed:
                radius = config_reader.get_value('Radius', 'curve_low_speed')
            elif 'medium' in vehicle_speed:
                radius = config_reader.get_value('Radius', 'curve_medium_speed')
            elif 'high' in vehicle_speed:
                radius = config_reader.get_value('Radius', 'curve_high_speed')
            else:
                raise Exception('Speed "{0}" not recognized in item {1}'.format(vehicle_speed, hara_item.item_id))
            try:
                self.road_radius = float(radius)
            except ValueError:
                raise ValueError("Invalid curve radius in config file, '{0}' in Radius section".format(radius))
        else:
            raise Exception('Route "{0}" not recognized'.format(route))

        if road_condition == '-' or 'dry' in road_condition or 'any' in road_condition:
            road_friction = config_reader.get_value('Road_friction', 'dry')
        elif 'wet' in road_condition:
            road_friction = config_reader.get_value('Road_friction', 'wet')
        elif 'icy' in road_condition or 'snow' in road_condition:
            road_friction = config_reader.get_value('Road_friction', 'icy')
            for i in range(len(self.vehicle_speed)):
                self.vehicle_speed[i] = min(self.vehicle_speed[i], 80)  # Vehicle speed is limited to 80 km/h on icy surfaces
        elif 'gravel' in road_condition:
            road_friction = config_reader.get_value('Road_friction', 'gravel')
        else:
            raise Exception('Road condition {0} not recognized in item {1}'.format(road_condition, hara_item.item_id))
        try:
            self.road_friction = float(road_friction)
        except ValueError:
            raise ValueError("Invalid road friction in config file, '{0}' in Road_friction section".format(road_friction))

        self.acceleration = None
        if (any(v != 0 for v in self.vehicle_speed)) and 'pressed' in brake_pedal:
            acceleration = config_reader.get_value('Driver', 'brake_pressed')
            if acceleration is not None:
                try:
                    self.acceleration = float(acceleration)
                except ValueError:
                    raise ValueError(
                        "Invalid driver input value in config file, '{0}' in Driver section".format(acceleration))

        self.faults = self.get_faults(hazard, engaged_gear)

    def get_faults(self, hazard: str, engaged_gear):
        """
Gets a list of the faults for the specified hazard
        :param hazard: Text of the hazard
        :param engaged_gear: Engaged gear (P, R, N or D)
        :return: Returns a list of the malfunctions for the hazard
        """
        faults = []
        if hazard is None:
            return []
        if engaged_gear is None or engaged_gear != 'R':
            direction = 1
        else:
            # Currently reverse driving is not supported by the VSM model. Therefore positive speeds are used in all scenarios.
            direction = 1
        front_wheel_angle = self.get_wheel_angle(self.road_radius, self.config_reader.get_float('Vehicle', 'wheelbase'))
        rear_wheel_angle = front_wheel_angle / 10

        if '[TQ1]' in hazard:
            faults.append(TorqueFault(torque_error_front=self.config_reader.get_float('Hazard_TQ', 'TQ1'),
                                      torque_error_rear=self.config_reader.get_float('Hazard_TQ', 'TQ1'),
                                      slew_rate=self.config_reader.get_float('Hazard_TQ', 'slew_rate')))
        elif '[TQ2]' in hazard:
            faults.append(TorqueFault(torque_error_front=self.config_reader.get_float('Hazard_TQ', 'TQ2'),
                                      torque_error_rear=self.config_reader.get_float('Hazard_TQ', 'TQ2'),
                                      slew_rate=self.config_reader.get_float('Hazard_TQ', 'slew_rate')))
            faults.append(TorqueFault(torque_error_front=self.config_reader.get_float('Hazard_TQ', 'TQ2'),
                                      slew_rate=self.config_reader.get_float('Hazard_TQ', 'slew_rate')))
            faults.append(TorqueFault(torque_error_rear=self.config_reader.get_float('Hazard_TQ', 'TQ2'),
                                      slew_rate=self.config_reader.get_float('Hazard_TQ', 'slew_rate')))
        elif '[TQ3]' in hazard:
            faults.append(TorqueFault(torque_error_front=direction * self.config_reader.get_float('Hazard_TQ', 'TQ3'),
                                      torque_error_rear=direction * self.config_reader.get_float('Hazard_TQ', 'TQ3'),
                                      slew_rate=self.config_reader.get_float('Hazard_TQ', 'slew_rate')))
        elif '[TQ4]' in hazard:
            faults.append(TorqueFault(torque_error_front=-1 * direction * self.config_reader.get_float('Hazard_TQ', 'TQ4'),
                                      torque_error_rear=-1 * direction * self.config_reader.get_float('Hazard_TQ', 'TQ4'),
                                      slew_rate=self.config_reader.get_float('Hazard_TQ', 'slew_rate')))
        elif '[TQ5]' in hazard:
            faults.append(TorqueFault(torque_error_front=-1 * direction * self.config_reader.get_float('Hazard_TQ', 'TQ5'),
                                      torque_error_rear=-1 * direction * self.config_reader.get_float('Hazard_TQ', 'TQ5'),
                                      slew_rate=self.config_reader.get_float('Hazard_TQ', 'slew_rate')))
        elif '[TQ6]' in hazard:
            faults.append(TorqueFault(torque_error_front=-1 * direction * self.config_reader.get_float('Hazard_TQ', 'TQ6'),
                                      slew_rate=self.config_reader.get_float('Hazard_TQ', 'slew_rate')))
            faults.append(TorqueFault(torque_error_rear=-1 * direction * self.config_reader.get_float('Hazard_TQ', 'TQ6'),
                                      slew_rate=self.config_reader.get_float('Hazard_TQ', 'slew_rate')))
            faults.append(TorqueFault(torque_error_front=-1 * direction * self.config_reader.get_float('Hazard_TQ', 'TQ6'),
                                      torque_error_rear=-1 * direction * self.config_reader.get_float('Hazard_TQ', 'TQ6'),
                                      slew_rate=self.config_reader.get_float('Hazard_TQ', 'slew_rate')))
            faults.append(TorqueFault(torque_error_front=direction * self.config_reader.get_float('Hazard_TQ', 'TQ6'),
                                      torque_error_rear=-1 * direction * self.config_reader.get_float('Hazard_TQ', 'TQ6'),
                                      slew_rate=self.config_reader.get_float('Hazard_TQ', 'slew_rate')))
            faults.append(TorqueFault(torque_error_front=-1 * direction * self.config_reader.get_float('Hazard_TQ', 'TQ6'),
                                      torque_error_rear=direction * self.config_reader.get_float('Hazard_TQ', 'TQ6'),
                                      slew_rate=self.config_reader.get_float('Hazard_TQ', 'slew_rate')))
        elif '[TQ7]' in hazard:
            faults.append(TorqueFault(torque_error_front=self.config_reader.get_float('Hazard_TQ', 'TQ7'),
                                      torque_error_rear=self.config_reader.get_float('Hazard_TQ', 'TQ7'),
                                      slew_rate=self.config_reader.get_float('Hazard_TQ', 'slew_rate')))
        elif '[TQ8]' in hazard:
            faults.append(TorqueFault(torque_error_front=self.config_reader.get_float('Hazard_TQ', 'TQ8'),
                                      torque_error_rear=self.config_reader.get_float('Hazard_TQ', 'TQ8'),
                                      slew_rate=self.config_reader.get_float('Hazard_TQ', 'slew_rate')))
        elif 'SUS' in hazard:
            if '[SUS1]' in hazard:
                # Too much steering goes until the physical limit of the steering system
                front_steering_limit = self.config_reader.get_float('Vehicle', 'front_steering_limit')
                additional_unintended_angle = front_steering_limit - front_wheel_angle
                faults.append(SteeringFault(additional_unintended_angle, self.config_reader.get_float('Hazard_SUS', 'slew_rate')))
            elif '[SUS2]' in hazard:
                faults.append(SteeringFault(-1 * self.config_reader.get_float('Hazard_SUS', 'SUS2') / 100 * front_wheel_angle, self.config_reader.get_float('Hazard_SUS', 'slew_rate')))
            elif '[SUS3]' in hazard:
                # Steering in opposite direction will be the same as the steering input from the driver but in the opposite direction.
                faults.append(SteeringFault(-2 * front_wheel_angle, self.config_reader.get_float('Hazard_SUS', 'slew_rate')))
        elif 'RAS' in hazard:
            if '[RAS1]' in hazard:
                # Too much steering goes until the physical limit of the steering system
                rear_steering_limit = self.config_reader.get_float('Vehicle', 'rear_steering_limit')
                additional_unintended_angle = rear_steering_limit - rear_wheel_angle
                faults.append(RearSteeringFault(additional_unintended_angle, self.config_reader.get_float('Hazard_RAS', 'slew_rate')))
            elif '[RAS3]' in hazard:
                faults.append(RearSteeringFault((-2 * rear_wheel_angle), self.config_reader.get_float('Hazard_RAS', 'slew_rate')))
        elif 'BS' in hazard:
            if '[BS1]' in hazard:
                faults.append(BrakingFault(self.config_reader.get_float('Hazard_BS', 'BS1')))
        elif 'PB' in hazard:
            faults.append(ParkBrakeFault(1))
        elif '[RHA1]' in hazard:
            faults.append(RideHeightFault.all_wheels(self.config_reader.get_float('Hazard_RHA', 'RHA1'),
                                                     slew_rate=self.config_reader.get_float('Hazard_RHA', 'slew_rate')))
        elif '[RHA2]' in hazard:
            faults.append(RideHeightFault.all_wheels(self.config_reader.get_float('Hazard_RHA', 'RHA2'),
                                                     slew_rate=self.config_reader.get_float('Hazard_RHA', 'slew_rate')))
        elif '[RHA3]' in hazard:
            faults.append(RideHeightFault(self.config_reader.get_float('Hazard_RHA', 'RHA3'),
                                          -1 * self.config_reader.get_float('Hazard_RHA', 'RHA3'),
                                          -1 * self.config_reader.get_float('Hazard_RHA', 'RHA3'),
                                          self.config_reader.get_float('Hazard_RHA', 'RHA3'),
                                          slew_rate=self.config_reader.get_float('Hazard_RHA', 'slew_rate')))
            faults.append(RideHeightFault(self.config_reader.get_float('Hazard_RHA', 'RHA3'),
                                          -1 * self.config_reader.get_float('Hazard_RHA', 'RHA3'),
                                          self.config_reader.get_float('Hazard_RHA', 'RHA3'),
                                          -1 * self.config_reader.get_float('Hazard_RHA', 'RHA3'),
                                          slew_rate=self.config_reader.get_float('Hazard_RHA', 'slew_rate')))
        return faults

    @staticmethod
    def get_wheel_angle(radius, wheelbase: float):
        """
Calculates the front wheel angles from the radius of the road
        :param radius: The road radius in meters
        :param wheelbase: The wheelbase of the vehicle in meters
        :return:
        """
        if isinstance(radius, str):
            if radius.lower() == 'straight':
                radius_float = 100.0  # For straight road driving, the steering angle of a 100 m curve is taken into consideration when defining too much steering malfunction
            else:
                raise ValueError("The specified road radius '{0}' is not valid".format(radius))
        else:
            try:
                radius_float = float(radius)
            except ValueError:
                raise ValueError("The specified road radius '{0}' is not valid".format(radius))
        wheel_angle_deg = math.asin(wheelbase / radius_float) * 180.0 / math.pi
        return wheel_angle_deg


class HaraReader:
    """
Reader for loading the HARA sheet and getting entries consecutively
    """

    def __init__(self, path, sheet_name, config_reader, header=0):
        if not os.path.isfile(path):
            raise Exception('Hara sheet was not found: {0}'.format(path))
        hara_workbook = openpyxl.load_workbook(path, data_only=True)
        self._sheet = hara_workbook[sheet_name]
        if self._sheet is None:
            raise Exception('Sheet {0} was not found in {1}'.format(sheet_name, path))
        if header < 0:
            raise Exception('Header size {} is invalid. It has to be greater or equal to 0'.format(header))
        self._header = header
        self._actual_row = header
        self._config_reader = config_reader
        self._load_indexes()

    def _load_indexes(self):
        self._idx_item = self._config_reader.get_int('Hara_Sheet', 'idx_item')
        self._idx_location = self._config_reader.get_int('Hara_Sheet', 'idx_location')
        self._idx_slope = self._config_reader.get_int('Hara_Sheet', 'idx_slope')
        self._idx_route = self._config_reader.get_int('Hara_Sheet', 'idx_route')
        self._idx_road_condition = self._config_reader.get_int('Hara_Sheet', 'idx_road_condition')
        self._idx_engaged_gear = self._config_reader.get_int('Hara_Sheet', 'idx_engaged_gear')
        self._idx_vehicle_speed = self._config_reader.get_int('Hara_Sheet', 'idx_vehicle_speed')
        self._idx_brake_pedal = self._config_reader.get_int('Hara_Sheet', 'idx_brake_pedal')
        self._idx_hazard = self._config_reader.get_int('Hara_Sheet', 'idx_hazard')
        self._idx_relevance = self._config_reader.get_int('Hara_Sheet', 'idx_relevance')
        self._idx_comment = self._config_reader.get_int('Hara_Sheet', 'idx_comment')

    def get_next_item(self):
        """
Getting the next entry from HARA
        :return: Returns a HaraItem containing all the info for the HARA entry
        """
        self._actual_row += 1
        item = self._sheet.cell(row=self._actual_row, column=self._idx_item).value
        location = self._sheet.cell(row=self._actual_row, column=self._idx_location).value
        slope = self._sheet.cell(row=self._actual_row, column=self._idx_slope).value
        route = self._sheet.cell(row=self._actual_row, column=self._idx_route).value
        road_condition = self._sheet.cell(row=self._actual_row, column=self._idx_road_condition).value
        engaged_gear = self._sheet.cell(row=self._actual_row, column=self._idx_engaged_gear).value
        vehicle_speed = self._sheet.cell(row=self._actual_row, column=self._idx_vehicle_speed).value
        brake_pedal = self._sheet.cell(row=self._actual_row, column=self._idx_brake_pedal).value
        hazard = self._sheet.cell(row=self._actual_row, column=self._idx_hazard).value
        relevance = self._sheet.cell(row=self._actual_row, column=self._idx_relevance).value == 'x'
        comment = self._sheet.cell(row=self._actual_row, column=self._idx_comment).value

        return HaraItem(item, location, slope, route, road_condition, engaged_gear, vehicle_speed, brake_pedal, hazard, relevance, comment)


class ScenarioWriter:
    """
Writer for writing the Scenario list to a file
    """

    def __init__(self, config_reader, template_path, sheet_name, header, mode):
        if not os.path.isfile(template_path):
            raise Exception('Scenario template was not found: {0}'.format(template_path))
        self._workbook = openpyxl.load_workbook(template_path)
        self._sheet = self._workbook[sheet_name]
        if self._sheet is None:
            raise Exception('Sheet {0} was not found in {1}'.format(sheet_name, template_path))
        if header < 0:
            raise Exception('Header size {} is invalid. It has to be greater or equal to 0'.format(header))
        self._header = header
        self._actual_row = header
        self._actual_test_run_id = 0
        self._config_reader = config_reader
        self._load_indexes()
        if mode.lower() == 'scenario_list':
            self._path = config_reader.get_value('Scenario_List', 'path')
        elif mode.lower() == 'ftti_list':
            self._path = config_reader.get_value('Scenario_List', 'ftti_path')
        elif mode.lower() == 'acceptance_list':
            self._path = config_reader.get_value('Scenario_List', 'acceptance_path')
        else:
            raise Exception("Mode '{0}' is not valid. Either use mode 'Scenario_List', 'FTTI_List' or 'Acceptance_List'".format(mode))
        self._mode = mode

    def _load_indexes(self):
        self._idx_hara_id = self._config_reader.get_int('Scenario_Template', 'idx_hara_id')
        self._idx_test_run_id = self._config_reader.get_int('Scenario_Template', 'idx_test_run_id')
        self._idx_constant_road_radius = self._config_reader.get_int('Scenario_Template', 'idx_constant_road_radius')
        self._idx_road_friction_coefficient = self._config_reader.get_int('Scenario_Template', 'idx_road_friction_coefficient')
        self._idx_road_gradient = self._config_reader.get_int('Scenario_Template', 'idx_road_gradient')
        self._idx_lateral_acceleration = self._config_reader.get_int('Scenario_Template', 'idx_lateral_acceleration')
        self._idx_friction_coefficient_exploitation = self._config_reader.get_int('Scenario_Template', 'idx_friction_coefficient_exploitation')
        self._idx_desired_vehicle_speed = self._config_reader.get_int('Scenario_Template', 'idx_desired_vehicle_speed')
        self._idx_acceleration = self._config_reader.get_int('Scenario_Template', 'idx_acceleration')

        self._idx_steering_front_angle = self._config_reader.get_int('Scenario_Template', 'idx_steering_front_angle')
        self._idx_steering_front_slew_rate = self._config_reader.get_int('Scenario_Template', 'idx_steering_front_slew_rate')
        self._idx_steering_rear_angle = self._config_reader.get_int('Scenario_Template', 'idx_steering_rear_angle')
        self._idx_steering_rear_slew_rate = self._config_reader.get_int('Scenario_Template', 'idx_steering_rear_slew_rate')
        self._idx_torque_front_axle = self._config_reader.get_int('Scenario_Template', 'idx_torque_front_axle')
        self._idx_torque_rear_axle = self._config_reader.get_int('Scenario_Template', 'idx_torque_rear_axle')
        self._idx_torque_slew_rate = self._config_reader.get_int('Scenario_Template', 'idx_torque_slew_rate')
        self._idx_ride_height_front_left = self._config_reader.get_int('Scenario_Template', 'idx_ride_height_front_left')
        self._idx_ride_height_front_right = self._config_reader.get_int('Scenario_Template', 'idx_ride_height_front_right')
        self._idx_ride_height_rear_left = self._config_reader.get_int('Scenario_Template', 'idx_ride_height_rear_left')
        self._idx_ride_height_rear_right = self._config_reader.get_int('Scenario_Template', 'idx_ride_height_rear_right')
        self._idx_ride_height_slew_rate = self._config_reader.get_int('Scenario_Template', 'idx_ride_height_slew_rate')
        self._idx_unintended_braking_torque = self._config_reader.get_int('Scenario_Template', 'idx_unintended_braking_torque')

        self._idx_very_slow_steering = self._config_reader.get_int('Scenario_Template', 'idx_very_slow_steering')
        self._idx_slow_steering = self._config_reader.get_int('Scenario_Template', 'idx_slow_steering')
        self._idx_braking = self._config_reader.get_int('Scenario_Template', 'idx_braking')

        self._idx_ftti = self._config_reader.get_int('Scenario_Template', 'idx_ftti')

    def clear_columns(self, idx_first_column):
        i_column = idx_first_column
        while True:
            if self._sheet.cell(row=self._header, column=i_column).value is None:
                break
            rowNum = len(self._sheet['A'])
            for iRow in range(1, rowNum + 1):
                try:
                    self._sheet.cell(row=iRow, column=i_column).value = None
                    cell_title = self._sheet.cell(iRow, i_column)
                    cell_title.fill = openpyxl.styles.PatternFill()
                except:
                    pass
            i_column += 1

    def delete_columns(self, idx_columns):
        for idx_column in idx_columns:
            self._sheet.unmerge_cells(start_row=1, start_column=idx_column, end_row=4, end_column=idx_column)
            self._sheet.delete_cols(idx_column, 1)

    def write_cell(self, idx_row, idx_col, value):
        self._sheet.cell(row=idx_row, column=idx_col).value = value

    def write(self, hara_item, scenario):
        """
 Method to deal with the writing of scenarios containing multiple faults and reactions
        :param hara_item: HARA entry
        :param scenario: Scenario
        """
        for speed in scenario.vehicle_speed:
            for fault in scenario.faults:
                reactions = self._get_reactions(fault, scenario)
                for reaction in reactions:
                    self._write_line(hara_item, scenario, speed, fault, reaction)

    def _write_line(self, hara_item, scenario, vehicle_speed, fault, reaction):
        """
Method to deal with the writing of scenarios with a single fault but multiple reactions
        :param hara_item: HARA entry
        :param scenario: Scenario
        :param vehicle_speed: Vehicle speed
        :param fault: A single malfunction
        :param reaction: Either None, one reaction or a list of reactions
        """
        self._actual_test_run_id += 1

        fault_level = [1]
        if self._mode.lower() == 'ftti_list' or self._mode.lower() == 'acceptance_list':
            if hara_item.comment is None:
                return
            target_test_run_id = int(hara_item.comment)
            if target_test_run_id != self._actual_test_run_id:
                return

            if self._mode.lower() == 'ftti_list':
                if '[TQ1]' in hara_item.hazard.upper():
                    ftti = [100, 200, 300, 400, 500]
                elif '[TQ2]' in hara_item.hazard.upper():
                    ftti = [100, 200, 300, 400, 500]
                elif '[TQ3]' in hara_item.hazard.upper():
                    ftti = [75, 150, 225, 300, 375]
                elif '[TQ4]' in hara_item.hazard.upper():
                    ftti = [75, 150, 225, 300, 375]
                elif '[TQ5]' in hara_item.hazard.upper():
                    ftti = [75, 150, 225, 300, 375]
                elif '[TQ6]' in hara_item.hazard.upper():
                    ftti = [75, 150, 225, 300, 375]
                elif '[SUS1]' in hara_item.hazard.upper():
                    ftti = [5, 10, 15, 20, 25]
                elif '[SUS2]' in hara_item.hazard.upper():
                    ftti = [10, 20, 30, 40, 50]
                elif '[SUS3]' in hara_item.hazard.upper():
                    ftti = [5, 10, 15, 20, 25]
                elif '[RAS1]' in hara_item.hazard.upper():
                    ftti = [5, 10, 15, 20, 25]
                elif '[RAS3]' in hara_item.hazard.upper():
                    ftti = [10, 20, 30, 40, 50]
                elif '[BS1]' in hara_item.hazard.upper():
                    ftti = [75, 150, 225, 300, 375]
                elif '[BS2]' in hara_item.hazard.upper():
                    ftti = [0, 5, 10, 15, 20]
                else:
                    raise Exception("The FTTI for {0} could not be determined. Hazard could not be recognized: {1}".format(hara_item.item_id, hara_item.hazard))
            else:
                ftti = [1000]
                fault_level = [.2, .4, .6, .8, 1]

            reaction = []
            for i_ftti in range(len(ftti)):
                reaction.append([VerySlowSteeringReaction(0), SlowSteeringReaction(0), BrakingReaction(20), FaultTolerantTime(ftti[i_ftti])])
            # reaction = [[VerySlowSteeringReaction(0), SlowSteeringReaction(0), BrakingReaction(20), FaultTolerantTime(ftti[0])],
            #             [VerySlowSteeringReaction(0), SlowSteeringReaction(0), BrakingReaction(20), FaultTolerantTime(ftti[1])]]

        ftti_cnt = 1
        if isinstance(reaction, list) and isinstance(reaction[0], list):
            ftti_cnt = len(reaction)
        else:
            reaction = [reaction]

        for i_ftti in range(ftti_cnt):
            for level in fault_level:
                self._actual_row += 1

                loc_test_run_id = self._actual_row - self._header
                # if self._mode.lower() == 'ftti_list':
                #     loc_test_run_id += 0

                print('Status: Writing item #{0}'.format(self._actual_row - self._header))

                self._sheet.cell(row=self._actual_row, column=self._idx_hara_id).value = hara_item.item_id
                self._sheet.cell(row=self._actual_row, column=self._idx_test_run_id).value = '%05d' % loc_test_run_id
                self._sheet.cell(row=self._actual_row, column=self._idx_constant_road_radius).value = scenario.road_radius
                self._sheet.cell(row=self._actual_row, column=self._idx_road_friction_coefficient).value = scenario.road_friction
                self._sheet.cell(row=self._actual_row, column=self._idx_road_gradient).value = scenario.road_gradient
                self._sheet.cell(row=self._actual_row, column=self._idx_lateral_acceleration).value = '=IF(ISNUMBER(C{0}),(H{0}/3.6)^2/C{0},"-")'.format(self._actual_row)
                self._sheet.cell(row=self._actual_row, column=self._idx_friction_coefficient_exploitation).value = '=IF(ISNUMBER(C{0}), F{0}/D{0}*100/9.81, "-")'.format(self._actual_row)
                self._sheet.cell(row=self._actual_row, column=self._idx_desired_vehicle_speed).value = vehicle_speed
                self._sheet.cell(row=self._actual_row, column=self._idx_acceleration).value = scenario.acceleration

                if fault is not None:
                    self._write_fault(fault, level)
                    # if mode.lower() == 'ftti_list':
                    #     self._write_reaction(BrakingReaction(99))
                    #     self._write_reaction(FaultTolerantTime(69))
                    # else:
                    if reaction[i_ftti] is not None:
                        if isinstance(reaction[i_ftti], list):
                            for single_reaction in reaction[i_ftti]:
                                self._write_reaction(single_reaction)
                        else:
                            self._write_reaction(reaction[i_ftti])

    def _write_reaction(self, reaction):
        if type(reaction) is BrakingReaction:
            self._sheet.cell(row=self._actual_row, column=self._idx_braking).value = reaction.braking
        elif type(reaction) is VerySlowSteeringReaction:
            self._sheet.cell(row=self._actual_row, column=self._idx_very_slow_steering).value = reaction.steering_rate_limit
        elif type(reaction) is SlowSteeringReaction:
            self._sheet.cell(row=self._actual_row, column=self._idx_slow_steering).value = reaction.steering_rate_limit
        elif type(reaction) is FaultTolerantTime:
            self._sheet.cell(row=self._actual_row, column=self._idx_ftti).value = reaction.ftti
        else:
            raise Exception("Type '{}' of the specified reaction is not valid".format(type(reaction)))

    def _write_fault(self, fault, level=1):
        if type(fault) is SteeringFault:
            self._sheet.cell(row=self._actual_row, column=self._idx_steering_front_angle).value = fault.additional_unintended_angle * level if fault.additional_unintended_angle is not None else fault.additional_unintended_angle
            self._sheet.cell(row=self._actual_row, column=self._idx_steering_front_slew_rate).value = fault.slew_rate
        elif type(fault) is RearSteeringFault:
            self._sheet.cell(row=self._actual_row, column=self._idx_steering_rear_angle).value = fault.additional_unintended_angle * level if fault.additional_unintended_angle is not None else fault.additional_unintended_angle
            self._sheet.cell(row=self._actual_row, column=self._idx_steering_rear_slew_rate).value = fault.slew_rate
        elif type(fault) is TorqueFault:
            self._sheet.cell(row=self._actual_row, column=self._idx_torque_front_axle).value = fault.torque_error_front * level if fault.torque_error_front is not None else fault.torque_error_front
            self._sheet.cell(row=self._actual_row, column=self._idx_torque_rear_axle).value = fault.torque_error_rear * level if fault.torque_error_rear is not None else fault.torque_error_rear
            self._sheet.cell(row=self._actual_row, column=self._idx_torque_slew_rate).value = fault.slew_rate
        elif type(fault) is RideHeightFault:
            self._sheet.cell(row=self._actual_row, column=self._idx_ride_height_front_left).value = fault.front_left * level if fault.front_left is not None else fault.front_left
            self._sheet.cell(row=self._actual_row, column=self._idx_ride_height_front_right).value = fault.front_right * level if fault.front_right is not None else fault.front_right
            self._sheet.cell(row=self._actual_row, column=self._idx_ride_height_rear_left).value = fault.rear_left * level if fault.rear_left is not None else fault.rear_left
            self._sheet.cell(row=self._actual_row, column=self._idx_ride_height_rear_right).value = fault.rear_right * level if fault.rear_right is not None else fault.rear_right
            self._sheet.cell(row=self._actual_row, column=self._idx_ride_height_slew_rate).value = fault.slew_rate
        elif type(fault) is BrakingFault:
            self._sheet.cell(row=self._actual_row, column=self._idx_unintended_braking_torque).value = fault.unintended_braking_torque * level if fault.unintended_braking_torque is not None else fault.unintended_braking_torque

    def _get_reactions(self, fault, scenario):
        """
Method to get the expected reactions based on the fault and road conditions
        :param fault: Malfunction
        :param scenario: Scenario
        :return: Returns all the expected reactions in a list
        """

        # No reaction is always considered as a separate testrun (for severity calculation):
        reactions = [[VerySlowSteeringReaction(0), SlowSteeringReaction(0), BrakingReaction(0)]]

        # The force of the braking reaction depends on the unintended acceleration level.
        # For higher unintended acceleration a stronger braking reaction is assumed.
        # The braking however is limited to a lower level on icy surfaces.

        if isinstance(fault, TorqueFault):
            if fault.get_overall_torque() > 100:
                braking_reaction = self._config_reader.get_float('Reaction', 'braking_torque_fault_high')
            elif fault.get_overall_torque() < 0:
                braking_reaction = 5
            else:
                braking_reaction = self._config_reader.get_float('Reaction', 'braking_torque_fault_low')
        else:
            braking_reaction = self._config_reader.get_float('Reaction', 'braking_normal')
        if scenario.road_friction <= self._config_reader.get_float('Road_friction', 'icy'):
            braking_reaction = min(braking_reaction, self._config_reader.get_float('Reaction', 'braking_low_friction'))

        # Braking without steering is a reaction that is expected always unless the fault is already leading to a high deceleration
        if (not isinstance(fault, TorqueFault) or fault.get_overall_torque() >= 0) and not isinstance(fault, BrakingFault):
            reactions.append([VerySlowSteeringReaction(0), SlowSteeringReaction(0), BrakingReaction(braking_reaction)])

        # Braking reaction together with steering correction is expected always except when driving on a straight road with a high friction
        friction_limit = self._config_reader.get_float('Road_friction', 'gravel')
        if (not isinstance(scenario.road_radius, str)) or scenario.road_friction < friction_limit or (isinstance(fault, TorqueFault) and fault.losing_stability()):
            # Braking with very slow steering reaction:
            reactions.append([BrakingReaction(braking_reaction), VerySlowSteeringReaction(self._config_reader.get_float('Reaction', 'very_slow_steering'))])
            # Braking with slow steering reaction:
            reactions.append([BrakingReaction(braking_reaction), SlowSteeringReaction(self._config_reader.get_float('Reaction', 'slow_steering'))])
        return reactions

    def save(self):
        """
Formatting the sheet and saving it
        """
        path = self._path
        print('Status: Saving to {}...'.format(path))
        for iCol in range(1, self._sheet.max_column):
            font = copy.copy(self._sheet.cell(row=self._header + 1, column=iCol).font)
            alignment = copy.copy(self._sheet.cell(row=self._header + 1, column=iCol).alignment)
            number_format = self._sheet.cell(row=self._header + 1, column=iCol).number_format
            for iRow in range(self._header + 2, self._actual_row + 1):
                self._sheet.cell(row=iRow, column=iCol).font = font
                self._sheet.cell(row=iRow, column=iCol).alignment = alignment
                self._sheet.cell(row=iRow, column=iCol).number_format = number_format
        if self._mode.lower() == 'ftti_list' or self._mode.lower() == 'acceptance_list':
            self._sheet.column_dimensions['Z'].hidden = True
            self._sheet.column_dimensions['AA'].hidden = False
            self._sheet.column_dimensions['AB'].hidden = False
            self._sheet.column_dimensions['AC'].hidden = True
            self._sheet.column_dimensions['AD'].hidden = True
            self._sheet.column_dimensions['AE'].hidden = True
            self._sheet.column_dimensions['AF'].hidden = True
            self._sheet.column_dimensions['AG'].hidden = True
            self._sheet.column_dimensions['AH'].hidden = True
            self._sheet.column_dimensions['AI'].hidden = False

        self._workbook.save(path)
        print('Status: Done')


if __name__ == '__main__':
    print('Status: Started')
    main('Scenario_List')
    # main('FTTI_List')
    # main('Acceptance_List')
