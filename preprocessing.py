"""
Generates a list of scenarios for the simulation using the HARA sheet as input.
"""
import copy
from dataclasses import dataclass
import os

import openpyxl
import openpyxl.styles

from packages.config import Config


def preprocessing(mode):
    """
Generates a list of scenarios for the simulation using the HARA sheet as input.
    """

    print('Status: Started')

    config_path = 'config.ini'
    config = Config(config_path)
    hara = Hara(config)
    scenario_list = ScenarioList(config, mode)

    for hazardous_event in hara.hazardous_events():
        if not hazardous_event.relevant:
            continue
        scenario = Scenario(config, hazardous_event)
        scenario_list.write(hazardous_event, scenario)
    scenario_list.save()

    print('Status: Done')


class TorqueFault:
    """
    E-motor torque malfunction
    """

    def __init__(self, torque_error_front=None, torque_error_rear=None, slew_rate=None):
        self.torque_error_front = torque_error_front
        self.torque_error_rear = torque_error_rear
        self.slew_rate = slew_rate

    def get_overall_torque(self):
        """
        Gets the overall torque for all wheels
        :return: Returns the overall torque by summing the front and rear e-motor torque
        """
        torque_front = self.torque_error_front if isinstance(self.torque_error_front, float) else 0.0
        torque_rear = self.torque_error_rear if isinstance(self.torque_error_rear, float) else 0.0
        overall_torque = torque_front + torque_rear
        return overall_torque

    def losing_stability(self) -> bool:
        """
        Checks the possibility of losing stability
        :return: Returns true if there is a possibility for losing stability
        """
        if self.torque_error_rear is not None:
            if self.torque_error_rear < 0:
                return True
        if self.torque_error_front is not None:
            if self.torque_error_front < 0:
                return True
        if self.torque_error_front is not None and self.torque_error_rear is not None:
            if self.torque_error_front < 0 and self.torque_error_rear > 0:  # pylint: disable=chained-comparison
                return True
            if self.torque_error_front > 0 and self.torque_error_rear < 0:  # pylint: disable=chained-comparison
                return True
        return False


@dataclass
class VerySlowSteeringReaction:
    """
    Applied steering reaction by the driver in degrees per second
    """
    steering_rate_limit: float


@dataclass
class SlowSteeringReaction:
    """
    Applied steering reaction by the driver in degrees per second
    """
    steering_rate_limit: float


@dataclass
class BrakingReaction:
    """
    Applied braking reaction by the driver in percentage
    """
    braking: float


@dataclass
class FaultTolerantTime:
    """
    Fault duration in milliseconds for the determination of the Fault Tolerant Time Interval (FTTI)
    """
    ftti: float


@dataclass
class HazardousEvent:  # pylint: disable=too-many-instance-attributes
    """
    Type containing all information for a Hazardous Event
    """
    identifier: str
    location: str
    slope: str
    route: str
    road_condition: str
    engaged_gear: str
    vehicle_speed: str
    brake_pedal: str
    maneuver: str
    hazard: str
    relevant: bool
    comment: str


class Scenario:  # pylint: disable=too-few-public-methods
    """
    Converts a Hazardous event to a Scenario (using the config settings)
    """

    def __init__(self, config, hazardous_event):
        self._config = config
        self._hazardous_event = hazardous_event
        slope = hazardous_event.slope.lower()
        route = hazardous_event.route.lower()
        road_condition = hazardous_event.road_condition.lower()
        engaged_gear = hazardous_event.engaged_gear.lower()
        vehicle_speed = hazardous_event.vehicle_speed.lower()
        brake_pedal = hazardous_event.brake_pedal.lower()
        maneuver = hazardous_event.maneuver.lower()

        self.road_gradient = self._get_road_gradient(slope)
        self.vehicle_speed = self._get_vehicle_speed(vehicle_speed, engaged_gear)
        self.road_radius = self._get_road_radius(route, vehicle_speed)
        self.road_friction = self._get_road_friction(road_condition)
        self.acceleration = self._get_acceleration(brake_pedal, maneuver)
        self.faults = self._get_faults(engaged_gear)

    def _get_road_gradient(self, slope):
        if slope == '-' or any(_ in slope for _ in ['any', 'flat']):
            # TODO: remove 'any' from the script, specify correctly the slope in the HARA
            road_gradient = self._config.get_entry('Slope', 'flat')
        elif 'slight' in slope:
            road_gradient = self._config.get_entry('Slope', 'slight_slope')
        elif 'downhill' in slope:
            road_gradient = self._config.get_entry('Slope', 'downhill')
        elif 'uphill' in slope:
            road_gradient = self._config.get_entry('Slope', 'uphill')
        else:
            raise KeyError(f"Slope {slope} not recognized in hazardous event {self._hazardous_event.identifier}")
        try:
            return float(road_gradient)
        except ValueError as exc:
            raise ValueError(f"Invalid road gradient '{road_gradient}' in config file, in Slope section") from exc

    def _get_vehicle_speed(self, vehicle_speed, engaged_gear):
        if vehicle_speed == '-' or any(_ in vehicle_speed for _ in ['any', 'stand']):
            # TODO: remove 'any' from the script, specify correctly the speed in the HARA
            speed_list_text = self._config.get_entry('Speed', 'standstill')
        elif 'very low' in vehicle_speed:
            speed_list_text = self._config.get_entry('Speed', 'very_low')
        elif 'low' in vehicle_speed:
            speed_list_text = self._config.get_entry('Speed', 'low')
        elif 'medium' in vehicle_speed:
            speed_list_text = self._config.get_entry('Speed', 'medium')
        elif 'high' in vehicle_speed:
            speed_list_text = self._config.get_entry('Speed', 'high')
        else:
            raise KeyError(f"Speed '{vehicle_speed}' not recognized "
                           f"in hazardous event {self._hazardous_event.identifier}")
        speed_list = speed_list_text.strip('[').strip(']').split(',')
        speed = [.0] * len(speed_list)
        for i, _ in enumerate(speed_list):
            try:
                speed[i] = float(speed_list[i]) if engaged_gear != 'r' else 1 * float(speed_list[i])
            except ValueError as exc:
                raise ValueError(f"Invalid speed thresholds in config file, "
                                 f"'{speed_list_text}' in Speed section") from exc
        return speed

    def _get_road_radius(self, route, vehicle_speed):
        if route == '-' or any(_ in route for _ in ['any', 'straight']):
            # TODO: remove 'any' from the script, specify correctly the route in the HARA
            road_radius = 'straight'
        elif 'curve' in route:
            if 'very_low' in vehicle_speed:
                radius = self._config.get_entry('Radius', 'curve_very_low_speed')
            elif vehicle_speed == '-' or any(_ in vehicle_speed for _ in ['any', 'stand', 'low']):
                # TODO: remove 'any' from the script, specify correctly the speed in the HARA
                radius = self._config.get_entry('Radius', 'curve_low_speed')
            elif 'medium' in vehicle_speed:
                radius = self._config.get_entry('Radius', 'curve_medium_speed')
            elif 'high' in vehicle_speed:
                radius = self._config.get_entry('Radius', 'curve_high_speed')
            else:
                raise KeyError(f"Speed '{vehicle_speed}' not recognized "
                               f"in hazardous event {self._hazardous_event.identifier}")
            radius_list = radius.strip('[').strip(']').split(',')
            road_radius = [.0] * len(radius_list)
            for i, _ in enumerate(radius_list):
                try:
                    road_radius[i] = float(radius_list[i])
                except ValueError as exc:
                    raise ValueError(f"Invalid curve radius in config file, '{radius}' in Radius section") from exc
            if len(road_radius) != len(self.vehicle_speed):
                raise ValueError("Invalid curve radius specification in config file, "
                                 "the number of radius specified has to match the number of speeds defined.")
        else:
            raise KeyError(f"Route '{route}' not recognized")
        return road_radius

    def _get_road_friction(self, road_condition):
        if road_condition == '-' or any(_ in road_condition for _ in ['any', 'dry']):
            road_friction_text = self._config.get_entry('Road_friction', 'dry')
        elif 'wet' in road_condition:
            road_friction_text = self._config.get_entry('Road_friction', 'wet')
        elif 'icy' in road_condition or 'snow' in road_condition:
            road_friction_text = self._config.get_entry('Road_friction', 'icy')
            for i, _ in enumerate(self.vehicle_speed):
                self.vehicle_speed[i] = min(self.vehicle_speed[i], 80)
        elif 'gravel' in road_condition:
            road_friction_text = self._config.get_entry('Road_friction', 'gravel')
        elif 'mu-split' in road_condition:
            road_friction_text = self._config.get_entry('Road_friction', 'mu-split')
        else:
            raise KeyError(f"Road condition {road_condition} not recognized "
                           f"in hazardous event {self._hazardous_event.identifier}")

        if 'mu-split' not in road_condition:
            try:
                road_friction = float(road_friction_text)
            except ValueError as exc:
                raise ValueError(f"Invalid road friction in config file, "
                                 f"'{road_friction_text}' in Road_friction section") from exc
        else:
            road_friction = road_friction_text

        return road_friction

    def _get_acceleration(self, brake_pedal, maneuver):
        if any(v != 0 for v in self.vehicle_speed):
            if 'pressed' in brake_pedal:
                acceleration_text = self._config.get_entry('Driver', 'brake_pressed')
            elif 'overtaking' in maneuver:
                acceleration_text = self._config.get_entry('Driver', 'overtaking')
            else:
                acceleration_text = 0
            try:
                acceleration = float(acceleration_text)
            except ValueError as exc:
                raise ValueError(f"Invalid driver input value in config file, "
                                 f"'{acceleration_text}' in Driver section") from exc
        else:
            acceleration = None

        return acceleration

    def _get_faults(self, engaged_gear):
        faults = []
        if self._hazardous_event.hazard is None:
            return []
        if engaged_gear is None or engaged_gear != 'R':
            direction = 1
        else:
            direction = 1

        if '[TQ1]' in self._hazardous_event.hazard:
            faults.append(TorqueFault(torque_error_front=self._config.get_float('Hazard_TQ', 'TQ1'),
                                      torque_error_rear=self._config.get_float('Hazard_TQ', 'TQ1'),
                                      slew_rate=self._config.get_float('Hazard_TQ', 'slew_rate')))
        elif '[TQ2]' in self._hazardous_event.hazard:
            faults.append(TorqueFault(torque_error_front=self._config.get_float('Hazard_TQ', 'TQ2'),
                                      torque_error_rear=self._config.get_float('Hazard_TQ', 'TQ2'),
                                      slew_rate=self._config.get_float('Hazard_TQ', 'slew_rate')))
            faults.append(TorqueFault(torque_error_front=self._config.get_float('Hazard_TQ', 'TQ2'),
                                      slew_rate=self._config.get_float('Hazard_TQ', 'slew_rate')))
            faults.append(TorqueFault(torque_error_rear=self._config.get_float('Hazard_TQ', 'TQ2'),
                                      slew_rate=self._config.get_float('Hazard_TQ', 'slew_rate')))
        elif '[TQ3]' in self._hazardous_event.hazard:
            faults.append(TorqueFault(torque_error_front=direction * self._config.get_float('Hazard_TQ', 'TQ3'),
                                      torque_error_rear=direction * self._config.get_float('Hazard_TQ', 'TQ3'),
                                      slew_rate=self._config.get_float('Hazard_TQ', 'slew_rate')))
        elif '[TQ4]' in self._hazardous_event.hazard:
            faults.append(TorqueFault(torque_error_front=-1 * direction * self._config.get_float('Hazard_TQ', 'TQ4'),
                                      torque_error_rear=-1 * direction * self._config.get_float('Hazard_TQ', 'TQ4'),
                                      slew_rate=self._config.get_float('Hazard_TQ', 'slew_rate')))
        elif '[TQ5]' in self._hazardous_event.hazard:
            faults.append(TorqueFault(torque_error_front=-1 * direction * self._config.get_float('Hazard_TQ', 'TQ5'),
                                      torque_error_rear=-1 * direction * self._config.get_float('Hazard_TQ', 'TQ5'),
                                      slew_rate=self._config.get_float('Hazard_TQ', 'slew_rate')))
        elif '[TQ6]' in self._hazardous_event.hazard:
            faults.append(TorqueFault(torque_error_front=-1 * direction * self._config.get_float('Hazard_TQ', 'TQ6'),
                                      slew_rate=self._config.get_float('Hazard_TQ', 'slew_rate')))
            faults.append(TorqueFault(torque_error_rear=-1 * direction * self._config.get_float('Hazard_TQ', 'TQ6'),
                                      slew_rate=self._config.get_float('Hazard_TQ', 'slew_rate')))
            faults.append(TorqueFault(torque_error_front=-1 * direction * self._config.get_float('Hazard_TQ', 'TQ6'),
                                      torque_error_rear=-1 * direction * self._config.get_float('Hazard_TQ', 'TQ6'),
                                      slew_rate=self._config.get_float('Hazard_TQ', 'slew_rate')))
            faults.append(TorqueFault(torque_error_front=direction * self._config.get_float('Hazard_TQ', 'TQ6'),
                                      torque_error_rear=-1 * direction * self._config.get_float('Hazard_TQ', 'TQ6'),
                                      slew_rate=self._config.get_float('Hazard_TQ', 'slew_rate')))
            faults.append(TorqueFault(torque_error_front=-1 * direction * self._config.get_float('Hazard_TQ', 'TQ6'),
                                      torque_error_rear=direction * self._config.get_float('Hazard_TQ', 'TQ6'),
                                      slew_rate=self._config.get_float('Hazard_TQ', 'slew_rate')))
        elif '[TQ7]' in self._hazardous_event.hazard:
            faults.append(TorqueFault(torque_error_front=self._config.get_float('Hazard_TQ', 'TQ7'),
                                      torque_error_rear=self._config.get_float('Hazard_TQ', 'TQ7'),
                                      slew_rate=self._config.get_float('Hazard_TQ', 'slew_rate')))
        return faults


class Hara:
    """
    Loads the HARA sheet and gets the hazardous events
    """

    def __init__(self, config):
        self._config = config
        hara_path = self._config.get_entry('Hara_Sheet', 'path')
        sheet_name = self._config.get_entry('Hara_Sheet', 'sheet_name')
        if not os.path.exists(hara_path):
            raise FileNotFoundError(f"Hara sheet was not found: {hara_path}")
        hara_workbook = openpyxl.load_workbook(hara_path, data_only=True)
        try:
            self._sheet = hara_workbook[sheet_name]
        except KeyError as exc:
            raise KeyError(f"Sheet {sheet_name} was not found in {hara_path}") from exc
        header_size = self._config.get_int('Hara_Sheet', 'header_size')
        if header_size < 0:
            raise ValueError(f"Header size {header_size} is invalid. It has to be greater or equal to 0")
        self._current_row = header_size
        self._indexes = self.Indexes(config)

    def _read_current_row(self, idx_column):
        return self._sheet.cell(row=self._current_row, column=idx_column).value

    def hazardous_events(self):
        """
        Gets the hazardous events from the HARA
        :return: Returns a HazardousEvents containing all the info for the hazardous event
        """
        while True:
            self._current_row += 1
            hazardous_event_id = self._read_current_row(self._indexes.id)
            location = self._read_current_row(self._indexes.location)
            slope = self._read_current_row(self._indexes.slope)
            route = self._read_current_row(self._indexes.route)
            road_condition = self._read_current_row(self._indexes.road_condition)
            engaged_gear = self._read_current_row(self._indexes.engaged_gear)
            vehicle_speed = self._read_current_row(self._indexes.vehicle_speed)
            brake_pedal = self._read_current_row(self._indexes.brake_pedal)
            maneuver = self._read_current_row(self._indexes.maneuver)
            hazard = self._read_current_row(self._indexes.hazard)
            relevance = self._read_current_row(self._indexes.relevance) == 'x'
            comment = self._read_current_row(self._indexes.comment)

            hazardous_event = HazardousEvent(hazardous_event_id, location, slope, route, road_condition, engaged_gear,
                                             vehicle_speed, brake_pedal, maneuver, hazard, relevance, comment)

            if hazardous_event.identifier is not None:
                yield hazardous_event
            else:
                break

    class Indexes:  # pylint: disable=too-many-instance-attributes disable=too-few-public-methods
        """
        Loads the indexes for the columns of the HARA sheet
        """
        def __init__(self, config):
            self.id = config.get_int('Hara_Sheet', 'idx_id')
            self.location = config.get_int('Hara_Sheet', 'idx_location')
            self.slope = config.get_int('Hara_Sheet', 'idx_slope')
            self.route = config.get_int('Hara_Sheet', 'idx_route')
            self.road_condition = config.get_int('Hara_Sheet', 'idx_road_condition')
            self.engaged_gear = config.get_int('Hara_Sheet', 'idx_engaged_gear')
            self.vehicle_speed = config.get_int('Hara_Sheet', 'idx_vehicle_speed')
            self.brake_pedal = config.get_int('Hara_Sheet', 'idx_brake_pedal')
            self.maneuver = config.get_int('Hara_Sheet', 'idx_maneuver')
            self.hazard = config.get_int('Hara_Sheet', 'idx_hazard')
            self.relevance = config.get_int('Hara_Sheet', 'idx_relevance')
            self.comment = config.get_int('Hara_Sheet', 'idx_comment')


class ScenarioList:
    """
    Generates the Scenario list to a file
    """

    def __init__(self, config, mode):
        self._config = config
        template_path = config.get_entry('Scenario_Template', 'path')
        sheet_name = config.get_entry('Scenario_Template', 'sheet_name')
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Scenario template was not found: {os.path.abspath(template_path)}")
        self._workbook = openpyxl.load_workbook(template_path)
        self._sheet = self._workbook[sheet_name]
        if self._sheet is None:
            raise KeyError(f"Sheet {sheet_name} was not found in {template_path}")
        header_size = self._config.get_int('Scenario_Template', 'header_size')
        if header_size < 0:
            raise ValueError(f"Header size {header_size} is invalid. It has to be greater or equal to 0")
        self._header_size = header_size
        self._current_row = header_size
        self._current_test_run_id = 0
        self._indexes = self.Indexes(self._config)
        if mode.lower() == 'scenario_list':
            self._path = self._config.get_entry('Scenario_List', 'path')
        elif mode.lower() == 'ftti_list':
            self._path = self._config.get_entry('Scenario_List', 'ftti_path')
        elif mode.lower() == 'acceptance_list':
            self._path = self._config.get_entry('Scenario_List', 'acceptance_path')
        else:
            raise ValueError(f"Mode '{mode}' is not valid. "
                             f"Either use mode 'Scenario_List', 'FTTI_List' or 'Acceptance_List'")
        self._mode = mode

    def _clear_columns(self, idx_first_column):
        i_column = idx_first_column
        while True:
            if self._sheet.cell(row=self._header_size, column=i_column).value is None:
                break
            column_length = len(self._sheet['A'])
            for i_row in range(self._header_size + 1, column_length + 1):
                self._sheet.cell(row=i_row, column=i_column).value = None
                cell_title = self._sheet.cell(i_row, i_column)
                cell_title.fill = openpyxl.styles.PatternFill()
            i_column += 1

    def _delete_columns(self, idx_columns):
        for idx_column in idx_columns:
            self._sheet.unmerge_cells(start_row=1, start_column=idx_column, end_row=4, end_column=idx_column)
            self._sheet.delete_cols(idx_column, 1)

    def _write_cell(self, idx_col, value):
        self._sheet.cell(row=self._current_row, column=idx_col).value = value

    def write(self, hazardous_event, scenario):
        """
        Method to deal with the writing of scenarios containing multiple faults and reactions
        :param hazardous_event: HARA entry
        :param scenario: Scenario
        """
        for i, speed in enumerate(scenario.vehicle_speed):
            radius = scenario.road_radius if isinstance(scenario.road_radius, str) else scenario.road_radius[i]
            for fault in scenario.faults:
                reactions = self._get_reactions(fault, scenario, hazardous_event)
                for reaction in reactions:
                    self._write_line(hazardous_event, scenario, speed, radius, fault, reaction)

    def _write_line(self, hazardous_event, scenario, vehicle_speed, road_radius, fault, reaction):
        """
        Method to deal with the writing of scenarios with a single fault but multiple reactions
        :param hazardous_event: HARA entry
        :param scenario: Scenario
        :param vehicle_speed: Vehicle speed
        :param road_radius: Road radius
        :param fault: A single malfunction
        :param reaction: Either None, one reaction or a list of reactions
        """
        self._current_test_run_id += 1

        if self._mode.lower() == 'ftti_list' or self._mode.lower() == 'acceptance_list':
            if hazardous_event.comment is None:
                return
            target_test_run_id = int(hazardous_event.comment)
            if target_test_run_id != self._current_test_run_id:
                return

            if self._mode.lower() == 'ftti_list':
                if '[TQ1]' in hazardous_event.hazard.upper():
                    ftti_list = [100, 200, 300, 400, 500]
                elif '[TQ2]' in hazardous_event.hazard.upper():
                    ftti_list = [100, 200, 300, 400, 500]
                elif '[TQ3]' in hazardous_event.hazard.upper():
                    ftti_list = [75, 150, 225, 300, 375]
                elif '[TQ4]' in hazardous_event.hazard.upper():
                    ftti_list = [75, 150, 225, 300, 375]
                elif '[TQ5]' in hazardous_event.hazard.upper():
                    ftti_list = [75, 150, 225, 300, 375]
                elif '[TQ6]' in hazardous_event.hazard.upper():
                    ftti_list = [75, 150, 225, 300, 375]
                else:
                    raise KeyError(f"The FTTI for {hazardous_event.identifier} could not be determined. "
                                   f"Hazard could not be recognized: {hazardous_event.hazard}")
            else:
                ftti_list = [1000]

            reaction = []
            for ftti in ftti_list:
                reaction.append([VerySlowSteeringReaction(0), SlowSteeringReaction(0),
                                 BrakingReaction(20), FaultTolerantTime(ftti)])

        ftti_cnt = 1
        if isinstance(reaction, list) and isinstance(reaction[0], list):
            ftti_cnt = len(reaction)
        else:
            reaction = [reaction]

        for i_ftti in range(ftti_cnt):
            self._current_row += 1

            loc_test_run_id = self._current_row - self._header_size

            print(f"Status: Writing scenario #{self._current_row - self._header_size}")

            self._write_cell(self._indexes.hara_id, hazardous_event.identifier)
            self._write_cell(self._indexes.test_run_id, f"{loc_test_run_id:05d}")
            self._write_cell(self._indexes.constant_road_radius, road_radius)
            self._write_cell(self._indexes.road_friction_coefficient, scenario.road_friction)
            self._write_cell(self._indexes.road_gradient, scenario.road_gradient)
            self._write_cell(self._indexes.lateral_acceleration,
                             f'=IF(ISNUMBER(C{self._current_row}), '
                             f'(H{self._current_row}/3.6)^2/C{self._current_row}, "-")')
            self._write_cell(self._indexes.friction_coefficient_exploitation,
                             f'=IF(ISNUMBER(C{self._current_row}), '
                             f'F{self._current_row}/D{self._current_row}*100/9.81, "-")')
            self._write_cell(self._indexes.desired_vehicle_speed, vehicle_speed)
            self._write_cell(self._indexes.acceleration, scenario.acceleration)

            if fault is not None:
                self._write_fault(fault)
                if reaction[i_ftti] is not None:
                    if isinstance(reaction[i_ftti], list):
                        for single_reaction in reaction[i_ftti]:
                            self._write_reaction(single_reaction)
                    else:
                        self._write_reaction(reaction[i_ftti])

    def _write_reaction(self, reaction):
        if isinstance(reaction, BrakingReaction):
            self._write_cell(self._indexes.braking, reaction.braking)
        elif isinstance(reaction, VerySlowSteeringReaction):
            self._write_cell(self._indexes.very_slow_steering, reaction.steering_rate_limit)
        elif isinstance(reaction, SlowSteeringReaction):
            self._write_cell(self._indexes.slow_steering, reaction.steering_rate_limit)
        elif isinstance(reaction, FaultTolerantTime):
            self._write_cell(self._indexes.ftti, reaction.ftti)
        else:
            raise TypeError(f"Type '{type(reaction)}' of the specified reaction is not valid")

    def _write_fault(self, fault):
        if isinstance(fault, TorqueFault):
            self._write_cell(self._indexes.torque_front_axle, fault.torque_error_front)
            self._write_cell(self._indexes.torque_rear_axle, fault.torque_error_rear)
            self._write_cell(self._indexes.torque_slew_rate, fault.slew_rate)

    def _get_reactions(self, fault, scenario, hazardous_event):
        """
        Method to get the expected reactions based on the fault and road conditions
        :param fault: Malfunction
        :param scenario: Scenario
        :param hazardous_event: Hazardous event
        :return: Returns all the expected reactions in a list
        """

        # No reaction is always considered as a separate testrun (for severity calculation):
        reactions = [[VerySlowSteeringReaction(0), SlowSteeringReaction(0), BrakingReaction(0)]]

        # The force of the braking reaction depends on the unintended acceleration level.
        # For higher unintended acceleration a stronger braking reaction is assumed.
        # The braking however is limited to a lower level on icy surfaces.

        if isinstance(fault, TorqueFault):
            if fault.get_overall_torque() > 100 or 'TQ4' in hazardous_event.hazard:
                braking_reaction = self._config.get_float('Reaction', 'braking_torque_fault_high')
            elif fault.get_overall_torque() < 0:
                braking_reaction = 5
            else:
                braking_reaction = self._config.get_float('Reaction', 'braking_torque_fault_low')
        else:
            braking_reaction = self._config.get_float('Reaction', 'braking_normal')
        if (isinstance(scenario.road_friction, float) and
                scenario.road_friction <= self._config.get_float('Road_friction', 'icy')):
            braking_reaction = min(braking_reaction, self._config.get_float('Reaction', 'braking_low_friction'))

        # Braking without steering is a reaction that is expected always
        reactions.append([VerySlowSteeringReaction(0), SlowSteeringReaction(0), BrakingReaction(braking_reaction)])

        # Braking reaction together with steering correction is expected always
        # except when driving on a straight road with a high friction
        friction_limit = self._config.get_float('Road_friction', 'gravel')
        if ((not isinstance(scenario.road_radius, str)) or
                (isinstance(scenario.road_friction, float) and scenario.road_friction < friction_limit) or
                (isinstance(fault, TorqueFault) and fault.losing_stability())):
            # Braking with very slow steering reaction:
            reactions.append([BrakingReaction(braking_reaction),
                              VerySlowSteeringReaction(self._config.get_float('Reaction', 'very_slow_steering'))])
            # Braking with slow steering reaction:
            reactions.append([BrakingReaction(braking_reaction),
                              SlowSteeringReaction(self._config.get_float('Reaction', 'slow_steering'))])
        return reactions

    def save(self):
        """
        Formatting the sheet and saving it
        """
        print(f"Status: Saving to {self._path}...")
        for i_col in range(1, self._sheet.max_column):
            font = copy.copy(self._sheet.cell(row=self._header_size + 1, column=i_col).font)
            alignment = copy.copy(self._sheet.cell(row=self._header_size + 1, column=i_col).alignment)
            number_format = self._sheet.cell(row=self._header_size + 1, column=i_col).number_format
            for i_row in range(self._header_size + 2, self._current_row + 1):
                self._sheet.cell(row=i_row, column=i_col).font = font
                self._sheet.cell(row=i_row, column=i_col).alignment = alignment
                self._sheet.cell(row=i_row, column=i_col).number_format = number_format
        if self._mode.lower() == 'ftti_list' or self._mode.lower() == 'acceptance_list':
            self._sheet.column_dimensions['Z'].hidden = True
            self._sheet.column_dimensions['AA'].hidden = False
            self._sheet.column_dimensions['AB'].hidden = False
            self._sheet.column_dimensions['AC'].hidden = True
            self._sheet.column_dimensions['AD'].hidden = True
            self._sheet.column_dimensions['AE'].hidden = True
            self._sheet.column_dimensions['AF'].hidden = True
            self._sheet.column_dimensions['AG'].hidden = True
            self._sheet.column_dimensions['AH'].hidden = True
            self._sheet.column_dimensions['AI'].hidden = False

        self._workbook.save(self._path)

    class Indexes:  # pylint: disable=too-many-instance-attributes disable=too-few-public-methods
        """
        Loads the indexes for the columns of the Scenario sheet
        """
        def __init__(self, config):
            self.hara_id = config.get_int('Scenario_Template', 'idx_hara_id')
            self.test_run_id = config.get_int('Scenario_Template', 'idx_test_run_id')
            self.constant_road_radius = config.get_int('Scenario_Template', 'idx_constant_road_radius')
            self.road_friction_coefficient = config.get_int('Scenario_Template', 'idx_road_friction_coefficient')
            self.road_gradient = config.get_int('Scenario_Template', 'idx_road_gradient')
            self.lateral_acceleration = config.get_int('Scenario_Template', 'idx_lateral_acceleration')
            self.friction_coefficient_exploitation = config.get_int('Scenario_Template',
                                                                    'idx_friction_coefficient_exploitation')
            self.desired_vehicle_speed = config.get_int('Scenario_Template', 'idx_desired_vehicle_speed')
            self.acceleration = config.get_int('Scenario_Template', 'idx_acceleration')
            self.torque_front_axle = config.get_int('Scenario_Template', 'idx_torque_front_axle')
            self.torque_rear_axle = config.get_int('Scenario_Template', 'idx_torque_rear_axle')
            self.torque_slew_rate = config.get_int('Scenario_Template', 'idx_torque_slew_rate')
            self.very_slow_steering = config.get_int('Scenario_Template', 'idx_very_slow_steering')
            self.slow_steering = config.get_int('Scenario_Template', 'idx_slow_steering')
            self.braking = config.get_int('Scenario_Template', 'idx_braking')
            self.ftti = config.get_int('Scenario_Template', 'idx_ftti')


if __name__ == '__main__':
    preprocessing('Scenario_List')
    # preprocessing('FTTI_List')
    # preprocessing('Acceptance_List')
