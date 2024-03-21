"""
Generates a list of scenarios for the simulation using the HARA sheet as input.
"""
import copy
import os
import openpyxl
import openpyxl.styles
from packages.config import Config


def main(mode):
    """
Generates a list of scenarios for the simulation using the HARA sheet as input.
    """

    config_path = 'config.ini'
    config = Config(config_path)

    hara_path = config.get_entry('Hara_Sheet', 'path')
    hara_sheet_name = config.get_entry('Hara_Sheet', 'sheet_name')
    hara = Hara(hara_path, hara_sheet_name, config)

    scenario_template_path = config.get_entry('Scenario_Template', 'path')
    scenario_template_sheet_name = config.get_entry('Scenario_Template', 'sheet_name')

    scenario_writer = ScenarioList(config, scenario_template_path, scenario_template_sheet_name, mode)

    for hazardous_event in hara.hazardous_events():
        if not hazardous_event.relevant:
            continue
        scenario = Scenario(config, hazardous_event)  # Converting the Hazardous Events to the Scenario list (using the config settings)
        scenario_writer.write(hazardous_event, scenario)  # Writing to the Scenario list
    scenario_writer.save()


class TorqueFault:
    """
E-motor torque malfunction
    """

    def __init__(self, torque_error_front=None, torque_error_rear=None, slew_rate=None):
        self.torque_error_front = torque_error_front
        self.torque_error_rear = torque_error_rear
        self.slew_rate = slew_rate

    def get_overall_torque(self):
        """
Calculates the overall torque
        :return: Returns the overall torque from the front and rear e-motors
        """
        torque_front = self.torque_error_front if isinstance(self.torque_error_front, float) else 0.0
        torque_rear = self.torque_error_rear if isinstance(self.torque_error_rear, float) else 0.0
        overall_torque = torque_front + torque_rear
        return overall_torque

    def losing_stability(self) -> bool:
        """
Checks the possibility of losing stability
        :return: Returns true if there is a possibility for losing stability
        """
        if self.torque_error_rear is not None:
            if self.torque_error_rear < 50:
                return True
        if self.torque_error_front is not None and self.torque_error_rear is not None:
            if self.torque_error_front < 0 and self.torque_error_rear > 0:
                return True
            if self.torque_error_front > 0 and self.torque_error_rear < 0:
                return True
        return False


class VerySlowSteeringReaction:
    """
Applied steering reaction in degrees per second
    """

    def __init__(self, steering_rate_limit):
        self.steering_rate_limit = steering_rate_limit


class SlowSteeringReaction:
    """
Applied steering reaction in degrees per second
    """

    def __init__(self, steering_rate_limit):
        self.steering_rate_limit = steering_rate_limit


class BrakingReaction:
    """
Applied braking force in percentage
    """

    def __init__(self, braking):
        self.braking = braking


class FaultTolerantTime:
    """
Applied braking force in percentage
    """

    def __init__(self, ftti):
        self.ftti = ftti


class HazardousEvent:
    """
Type containing the properties for a Hazardous Event
    """
    def __init__(self, hazardous_event_id, location, slope, route, road_condition, engaged_gear, vehicle_speed, brake_pedal, hazard, relevant, comment):
        self.id = hazardous_event_id
        self.location = location
        self.slope = slope
        self.route = route
        self.road_condition = road_condition
        self.engaged_gear = engaged_gear
        self.vehicle_speed = vehicle_speed
        self.brake_pedal = brake_pedal
        self.hazard = hazard
        self.relevant = relevant
        self.comment = comment


class Scenario:
    """
Converts a Hazardous event to a Scenario (using the config settings)
    """

    def __init__(self, config, hazardous_event):
        self.config = config
        slope = hazardous_event.slope.lower()
        route = hazardous_event.route.lower()
        road_condition = hazardous_event.road_condition.lower()
        engaged_gear = hazardous_event.engaged_gear.lower()
        vehicle_speed = hazardous_event.vehicle_speed.lower()
        brake_pedal = hazardous_event.brake_pedal.lower()
        hazard = hazardous_event.hazard

        if any(_ in slope for _ in ['any', '-', 'flat']):  # TODO: remove 'any' from the script, specify correctly the slope in the HARA
            road_gradient = config.get_entry('Slope', 'flat')
        elif 'slight' in slope:
            road_gradient = config.get_entry('Slope', 'slight_slope')
        elif 'down' in slope:
            road_gradient = config.get_entry('Slope', 'downhill')
        elif 'up' in slope:
            road_gradient = config.get_entry('Slope', 'uphill')
        else:
            raise Exception(f"Slope {slope} not recognized in hazardous event {hazardous_event.id}")
        try:
            self.road_gradient = float(road_gradient)
        except ValueError:
            raise ValueError(f"Invalid road gradient '{road_gradient}' in config file, in Slope section")

        if any(_ in vehicle_speed for _ in ['any', '-', 'stand']):  # TODO: remove 'any' from the script, specify correctly the speed in the HARA
            speed = config.get_entry('Speed', 'standstill')
        elif 'very low' in vehicle_speed:
            speed = config.get_entry('Speed', 'very_low')
        elif 'low' in vehicle_speed:
            speed = config.get_entry('Speed', 'low')
        elif 'medium' in vehicle_speed:
            speed = config.get_entry('Speed', 'medium')
        elif 'high' in vehicle_speed:
            speed = config.get_entry('Speed', 'high')
        else:
            raise Exception(f"Speed '{vehicle_speed}' not recognized in hazardous event {hazardous_event.id}")
        speed_list = speed.strip('[').strip(']').split(',')
        self.vehicle_speed = [.0] * len(speed_list)
        for i in range(len(speed_list)):
            try:
                # Currently reverse driving is not supported by the VSM model. Therefore, positive speeds are used in all scenarios.
                self.vehicle_speed[i] = float(speed_list[i]) if engaged_gear != 'r' else 1 * float(speed_list[i])
            except ValueError:
                raise ValueError(f"Invalid speed thresholds in config file, '{speed}' in Speed section")

        if route == '-' or 'straight' in route or 'any' in route:
            self.road_radius = 'straight'
        elif 'curve' in route:
            if vehicle_speed == '-' or 'stand' in vehicle_speed or 'low' in vehicle_speed or 'any' in vehicle_speed:  # TODO: remove 'any' from the script, specify correctly the speed in the HARA
                radius = config.get_entry('Radius', 'curve_low_speed')
            elif 'medium' in vehicle_speed:
                radius = config.get_entry('Radius', 'curve_medium_speed')
            elif 'high' in vehicle_speed:
                radius = config.get_entry('Radius', 'curve_high_speed')
            else:
                raise Exception(f"Speed '{vehicle_speed}' not recognized in hazardous event {hazardous_event.id}")
            try:
                self.road_radius = float(radius)
            except ValueError:
                raise ValueError(f"Invalid curve radius in config file, '{radius}' in Radius section")
        else:
            raise Exception(f"Route '{route}' not recognized")

        if road_condition == '-' or 'dry' in road_condition or 'any' in road_condition:
            road_friction = config.get_entry('Road_friction', 'dry')
        elif 'wet' in road_condition:
            road_friction = config.get_entry('Road_friction', 'wet')
        elif 'icy' in road_condition or 'snow' in road_condition:
            road_friction = config.get_entry('Road_friction', 'icy')
            for i in range(len(self.vehicle_speed)):
                self.vehicle_speed[i] = min(self.vehicle_speed[i], 80)  # Vehicle speed is limited to 80 km/h on icy surfaces
        elif 'gravel' in road_condition:
            road_friction = config.get_entry('Road_friction', 'gravel')
        else:
            raise Exception(f"Road condition {road_condition} not recognized in hazardous event {hazardous_event.id}")
        try:
            self.road_friction = float(road_friction)
        except ValueError:
            raise ValueError(f"Invalid road friction in config file, '{road_friction}' in Road_friction section")

        self.acceleration = None
        if (any(v != 0 for v in self.vehicle_speed)) and 'pressed' in brake_pedal:
            acceleration = config.get_entry('Driver', 'brake_pressed')
            if acceleration is not None:
                try:
                    self.acceleration = float(acceleration)
                except ValueError:
                    raise ValueError(f"Invalid driver input value in config file, '{acceleration}' in Driver section")

        self.faults = self.get_faults(hazard, engaged_gear)

    def get_faults(self, hazard: str, engaged_gear):
        """
Gets a list of the faults for the specified hazard
        :param hazard: Text of the hazard
        :param engaged_gear: Engaged gear (P, R, N or D)
        :return: Returns a list of the malfunctions for the hazard
        """
        faults = []
        if hazard is None:
            return []
        if engaged_gear is None or engaged_gear != 'R':
            direction = 1
        else:
            # Currently reverse driving is not supported by the VSM model. Therefore, positive speeds are used in all scenarios.
            direction = 1

        if '[TQ1]' in hazard:
            faults.append(TorqueFault(torque_error_front=self.config.get_float('Hazard_TQ', 'TQ1'),
                                      torque_error_rear=self.config.get_float('Hazard_TQ', 'TQ1'),
                                      slew_rate=self.config.get_float('Hazard_TQ', 'slew_rate')))
        elif '[TQ2]' in hazard:
            faults.append(TorqueFault(torque_error_front=self.config.get_float('Hazard_TQ', 'TQ2'),
                                      torque_error_rear=self.config.get_float('Hazard_TQ', 'TQ2'),
                                      slew_rate=self.config.get_float('Hazard_TQ', 'slew_rate')))
            faults.append(TorqueFault(torque_error_front=self.config.get_float('Hazard_TQ', 'TQ2'),
                                      slew_rate=self.config.get_float('Hazard_TQ', 'slew_rate')))
            faults.append(TorqueFault(torque_error_rear=self.config.get_float('Hazard_TQ', 'TQ2'),
                                      slew_rate=self.config.get_float('Hazard_TQ', 'slew_rate')))
        elif '[TQ3]' in hazard:
            faults.append(TorqueFault(torque_error_front=direction * self.config.get_float('Hazard_TQ', 'TQ3'),
                                      torque_error_rear=direction * self.config.get_float('Hazard_TQ', 'TQ3'),
                                      slew_rate=self.config.get_float('Hazard_TQ', 'slew_rate')))
        elif '[TQ4]' in hazard:
            faults.append(TorqueFault(torque_error_front=-1 * direction * self.config.get_float('Hazard_TQ', 'TQ4'),
                                      torque_error_rear=-1 * direction * self.config.get_float('Hazard_TQ', 'TQ4'),
                                      slew_rate=self.config.get_float('Hazard_TQ', 'slew_rate')))
        elif '[TQ5]' in hazard:
            faults.append(TorqueFault(torque_error_front=-1 * direction * self.config.get_float('Hazard_TQ', 'TQ5'),
                                      torque_error_rear=-1 * direction * self.config.get_float('Hazard_TQ', 'TQ5'),
                                      slew_rate=self.config.get_float('Hazard_TQ', 'slew_rate')))
        elif '[TQ6]' in hazard:
            faults.append(TorqueFault(torque_error_front=-1 * direction * self.config.get_float('Hazard_TQ', 'TQ6'),
                                      slew_rate=self.config.get_float('Hazard_TQ', 'slew_rate')))
            faults.append(TorqueFault(torque_error_rear=-1 * direction * self.config.get_float('Hazard_TQ', 'TQ6'),
                                      slew_rate=self.config.get_float('Hazard_TQ', 'slew_rate')))
            faults.append(TorqueFault(torque_error_front=-1 * direction * self.config.get_float('Hazard_TQ', 'TQ6'),
                                      torque_error_rear=-1 * direction * self.config.get_float('Hazard_TQ', 'TQ6'),
                                      slew_rate=self.config.get_float('Hazard_TQ', 'slew_rate')))
            faults.append(TorqueFault(torque_error_front=direction * self.config.get_float('Hazard_TQ', 'TQ6'),
                                      torque_error_rear=-1 * direction * self.config.get_float('Hazard_TQ', 'TQ6'),
                                      slew_rate=self.config.get_float('Hazard_TQ', 'slew_rate')))
            faults.append(TorqueFault(torque_error_front=-1 * direction * self.config.get_float('Hazard_TQ', 'TQ6'),
                                      torque_error_rear=direction * self.config.get_float('Hazard_TQ', 'TQ6'),
                                      slew_rate=self.config.get_float('Hazard_TQ', 'slew_rate')))
        elif '[TQ7]' in hazard:
            faults.append(TorqueFault(torque_error_front=self.config.get_float('Hazard_TQ', 'TQ7'),
                                      torque_error_rear=self.config.get_float('Hazard_TQ', 'TQ7'),
                                      slew_rate=self.config.get_float('Hazard_TQ', 'slew_rate')))
        elif '[TQ8]' in hazard:
            faults.append(TorqueFault(torque_error_front=self.config.get_float('Hazard_TQ', 'TQ8'),
                                      torque_error_rear=self.config.get_float('Hazard_TQ', 'TQ8'),
                                      slew_rate=self.config.get_float('Hazard_TQ', 'slew_rate')))
        return faults


class Hara:
    """
Loads the HARA sheet and gets the hazardous events
    """

    def __init__(self, path, sheet_name, config):
        if not os.path.exists(path):
            raise Exception(f"Hara sheet was not found: {path}")
        hara_workbook = openpyxl.load_workbook(path, data_only=True)
        self._sheet = hara_workbook[sheet_name]
        if self._sheet is None:
            raise Exception(f"Sheet {sheet_name} was not found in {path}")
        header_size = config.get_int('Hara_Sheet', 'header_size')
        if header_size < 0:
            raise Exception(f"Header size {header_size} is invalid. It has to be greater or equal to 0")
        self._current_row = header_size
        self._config = config
        self._load_indexes()

    def _load_indexes(self):
        self._idx_id = self._config.get_int('Hara_Sheet', 'idx_id')
        self._idx_location = self._config.get_int('Hara_Sheet', 'idx_location')
        self._idx_slope = self._config.get_int('Hara_Sheet', 'idx_slope')
        self._idx_route = self._config.get_int('Hara_Sheet', 'idx_route')
        self._idx_road_condition = self._config.get_int('Hara_Sheet', 'idx_road_condition')
        self._idx_engaged_gear = self._config.get_int('Hara_Sheet', 'idx_engaged_gear')
        self._idx_vehicle_speed = self._config.get_int('Hara_Sheet', 'idx_vehicle_speed')
        self._idx_brake_pedal = self._config.get_int('Hara_Sheet', 'idx_brake_pedal')
        self._idx_hazard = self._config.get_int('Hara_Sheet', 'idx_hazard')
        self._idx_relevance = self._config.get_int('Hara_Sheet', 'idx_relevance')
        self._idx_comment = self._config.get_int('Hara_Sheet', 'idx_comment')

    def _read_current_row(self, idx_column):
        return self._sheet.cell(row=self._current_row, column=idx_column).value

    def hazardous_events(self):
        """
Gets the hazardous events from the HARA
        :return: Returns a HazardousEvents containing all the info for the hazardous event
        """
        while True:
            self._current_row += 1
            hazardous_event_id = self._read_current_row(self._idx_id)
            location = self._read_current_row(self._idx_location)
            slope = self._read_current_row(self._idx_slope)
            route = self._read_current_row(self._idx_route)
            road_condition = self._read_current_row(self._idx_road_condition)
            engaged_gear = self._read_current_row(self._idx_engaged_gear)
            vehicle_speed = self._read_current_row(self._idx_vehicle_speed)
            brake_pedal = self._read_current_row(self._idx_brake_pedal)
            hazard = self._read_current_row(self._idx_hazard)
            relevance = self._read_current_row(self._idx_relevance) == 'x'
            comment = self._read_current_row(self._idx_comment)

            hazardous_event = HazardousEvent(hazardous_event_id, location, slope, route, road_condition, engaged_gear, vehicle_speed, brake_pedal, hazard, relevance, comment)

            if hazardous_event.id is not None:
                yield hazardous_event
            else:
                break


class ScenarioList:
    """
Generates the Scenario list to a file
    """

    def __init__(self, config, template_path, sheet_name, mode):
        if not os.path.exists(template_path):
            raise Exception(f"Scenario template was not found: {os.path.abspath(template_path)}")
        self._workbook = openpyxl.load_workbook(template_path)
        self._sheet = self._workbook[sheet_name]
        if self._sheet is None:
            raise Exception(f"Sheet {sheet_name} was not found in {template_path}")
        header_size = config.get_int('Scenario_Template', 'header_size')
        if header_size < 0:
            raise Exception(f"Header size {header_size} is invalid. It has to be greater or equal to 0")
        self._header_size = header_size
        self._current_row = header_size
        self._current_test_run_id = 0
        self._config = config
        self._load_indexes()
        if mode.lower() == 'scenario_list':
            self._path = config.get_entry('Scenario_List', 'path')
        elif mode.lower() == 'ftti_list':
            self._path = config.get_entry('Scenario_List', 'ftti_path')
        elif mode.lower() == 'acceptance_list':
            self._path = config.get_entry('Scenario_List', 'acceptance_path')
        else:
            raise Exception(f"Mode '{mode}' is not valid. Either use mode 'Scenario_List', 'FTTI_List' or 'Acceptance_List'")
        self._mode = mode

    def _load_indexes(self):
        self._idx_hara_id = self._config.get_int('Scenario_Template', 'idx_hara_id')
        self._idx_test_run_id = self._config.get_int('Scenario_Template', 'idx_test_run_id')
        self._idx_constant_road_radius = self._config.get_int('Scenario_Template', 'idx_constant_road_radius')
        self._idx_road_friction_coefficient = self._config.get_int('Scenario_Template', 'idx_road_friction_coefficient')
        self._idx_road_gradient = self._config.get_int('Scenario_Template', 'idx_road_gradient')
        self._idx_lateral_acceleration = self._config.get_int('Scenario_Template', 'idx_lateral_acceleration')
        self._idx_friction_coefficient_exploitation = self._config.get_int('Scenario_Template', 'idx_friction_coefficient_exploitation')
        self._idx_desired_vehicle_speed = self._config.get_int('Scenario_Template', 'idx_desired_vehicle_speed')
        self._idx_acceleration = self._config.get_int('Scenario_Template', 'idx_acceleration')

        self._idx_steering_front_angle = self._config.get_int('Scenario_Template', 'idx_steering_front_angle')
        self._idx_steering_front_slew_rate = self._config.get_int('Scenario_Template', 'idx_steering_front_slew_rate')
        self._idx_steering_rear_angle = self._config.get_int('Scenario_Template', 'idx_steering_rear_angle')
        self._idx_steering_rear_slew_rate = self._config.get_int('Scenario_Template', 'idx_steering_rear_slew_rate')
        self._idx_torque_front_axle = self._config.get_int('Scenario_Template', 'idx_torque_front_axle')
        self._idx_torque_rear_axle = self._config.get_int('Scenario_Template', 'idx_torque_rear_axle')
        self._idx_torque_slew_rate = self._config.get_int('Scenario_Template', 'idx_torque_slew_rate')
        self._idx_ride_height_front_left = self._config.get_int('Scenario_Template', 'idx_ride_height_front_left')
        self._idx_ride_height_front_right = self._config.get_int('Scenario_Template', 'idx_ride_height_front_right')
        self._idx_ride_height_rear_left = self._config.get_int('Scenario_Template', 'idx_ride_height_rear_left')
        self._idx_ride_height_rear_right = self._config.get_int('Scenario_Template', 'idx_ride_height_rear_right')
        self._idx_ride_height_slew_rate = self._config.get_int('Scenario_Template', 'idx_ride_height_slew_rate')
        self._idx_unintended_braking_torque = self._config.get_int('Scenario_Template', 'idx_unintended_braking_torque')

        self._idx_very_slow_steering = self._config.get_int('Scenario_Template', 'idx_very_slow_steering')
        self._idx_slow_steering = self._config.get_int('Scenario_Template', 'idx_slow_steering')
        self._idx_braking = self._config.get_int('Scenario_Template', 'idx_braking')

        self._idx_ftti = self._config.get_int('Scenario_Template', 'idx_ftti')

    def clear_columns(self, idx_first_column):
        i_column = idx_first_column
        while True:
            if self._sheet.cell(row=self._header_size, column=i_column).value is None:
                break
            column_length = len(self._sheet['A'])
            for i_row in range(self._header_size + 1, column_length + 1):
                try:
                    self._sheet.cell(row=i_row, column=i_column).value = None
                    cell_title = self._sheet.cell(i_row, i_column)
                    cell_title.fill = openpyxl.styles.PatternFill()
                except:
                    pass
            i_column += 1

    def delete_columns(self, idx_columns):
        for idx_column in idx_columns:
            self._sheet.unmerge_cells(start_row=1, start_column=idx_column, end_row=4, end_column=idx_column)
            self._sheet.delete_cols(idx_column, 1)

    def write_cell(self, idx_row, idx_col, value):
        self._sheet.cell(row=idx_row, column=idx_col).value = value

    def write(self, hazardous_event, scenario):
        """
 Method to deal with the writing of scenarios containing multiple faults and reactions
        :param hazardous_event: HARA entry
        :param scenario: Scenario
        """
        for speed in scenario.vehicle_speed:
            for fault in scenario.faults:
                reactions = self._get_reactions(fault, scenario)
                for reaction in reactions:
                    self._write_line(hazardous_event, scenario, speed, fault, reaction)

    def _write_line(self, hazardous_event, scenario, vehicle_speed, fault, reaction):
        """
Method to deal with the writing of scenarios with a single fault but multiple reactions
        :param hazardous_event: HARA entry
        :param scenario: Scenario
        :param vehicle_speed: Vehicle speed
        :param fault: A single malfunction
        :param reaction: Either None, one reaction or a list of reactions
        """
        self._current_test_run_id += 1

        fault_level = [1]
        if self._mode.lower() == 'ftti_list' or self._mode.lower() == 'acceptance_list':
            if hazardous_event.comment is None:
                return
            target_test_run_id = int(hazardous_event.comment)
            if target_test_run_id != self._current_test_run_id:
                return

            if self._mode.lower() == 'ftti_list':
                if '[TQ1]' in hazardous_event.hazard.upper():
                    ftti = [100, 200, 300, 400, 500]
                elif '[TQ2]' in hazardous_event.hazard.upper():
                    ftti = [100, 200, 300, 400, 500]
                elif '[TQ3]' in hazardous_event.hazard.upper():
                    ftti = [75, 150, 225, 300, 375]
                elif '[TQ4]' in hazardous_event.hazard.upper():
                    ftti = [75, 150, 225, 300, 375]
                elif '[TQ5]' in hazardous_event.hazard.upper():
                    ftti = [75, 150, 225, 300, 375]
                elif '[TQ6]' in hazardous_event.hazard.upper():
                    ftti = [75, 150, 225, 300, 375]
                else:
                    raise Exception(f"The FTTI for {hazardous_event.id} could not be determined. Hazard could not be recognized: {hazardous_event.hazard}")
            else:
                ftti = [1000]
                fault_level = [.2, .4, .6, .8, 1]

            reaction = []
            for i_ftti in range(len(ftti)):
                reaction.append([VerySlowSteeringReaction(0), SlowSteeringReaction(0), BrakingReaction(20), FaultTolerantTime(ftti[i_ftti])])
            # reaction = [[VerySlowSteeringReaction(0), SlowSteeringReaction(0), BrakingReaction(20), FaultTolerantTime(ftti[0])],
            #             [VerySlowSteeringReaction(0), SlowSteeringReaction(0), BrakingReaction(20), FaultTolerantTime(ftti[1])]]

        ftti_cnt = 1
        if isinstance(reaction, list) and isinstance(reaction[0], list):
            ftti_cnt = len(reaction)
        else:
            reaction = [reaction]

        for i_ftti in range(ftti_cnt):
            for level in fault_level:
                self._current_row += 1

                loc_test_run_id = self._current_row - self._header_size
                # if self._mode.lower() == 'ftti_list':
                #     loc_test_run_id += 0

                print(f"Status: Writing item #{self._current_row - self._header_size}")

                self._sheet.cell(row=self._current_row, column=self._idx_hara_id).value = hazardous_event.id
                self._sheet.cell(row=self._current_row, column=self._idx_test_run_id).value = '%05d' % loc_test_run_id
                self._sheet.cell(row=self._current_row, column=self._idx_constant_road_radius).value = scenario.road_radius
                self._sheet.cell(row=self._current_row, column=self._idx_road_friction_coefficient).value = scenario.road_friction
                self._sheet.cell(row=self._current_row, column=self._idx_road_gradient).value = scenario.road_gradient
                self._sheet.cell(row=self._current_row, column=self._idx_lateral_acceleration).value = '=IF(ISNUMBER(C{0}),(H{0}/3.6)^2/C{0},"-")'.format(self._current_row)
                self._sheet.cell(row=self._current_row, column=self._idx_friction_coefficient_exploitation).value = '=IF(ISNUMBER(C{0}), F{0}/D{0}*100/9.81, "-")'.format(self._current_row)
                self._sheet.cell(row=self._current_row, column=self._idx_desired_vehicle_speed).value = vehicle_speed
                self._sheet.cell(row=self._current_row, column=self._idx_acceleration).value = scenario.acceleration

                if fault is not None:
                    self._write_fault(fault, level)
                    # if mode.lower() == 'ftti_list':
                    #     self._write_reaction(BrakingReaction(99))
                    #     self._write_reaction(FaultTolerantTime(69))
                    # else:
                    if reaction[i_ftti] is not None:
                        if isinstance(reaction[i_ftti], list):
                            for single_reaction in reaction[i_ftti]:
                                self._write_reaction(single_reaction)
                        else:
                            self._write_reaction(reaction[i_ftti])

    def _write_reaction(self, reaction):
        if type(reaction) is BrakingReaction:
            self._sheet.cell(row=self._current_row, column=self._idx_braking).value = reaction.braking
        elif type(reaction) is VerySlowSteeringReaction:
            self._sheet.cell(row=self._current_row, column=self._idx_very_slow_steering).value = reaction.steering_rate_limit
        elif type(reaction) is SlowSteeringReaction:
            self._sheet.cell(row=self._current_row, column=self._idx_slow_steering).value = reaction.steering_rate_limit
        elif type(reaction) is FaultTolerantTime:
            self._sheet.cell(row=self._current_row, column=self._idx_ftti).value = reaction.ftti
        else:
            raise Exception(f"Type '{type(reaction)}' of the specified reaction is not valid")

    def _write_fault(self, fault, level=1):
        if type(fault) is TorqueFault:
            self._sheet.cell(row=self._current_row, column=self._idx_torque_front_axle).value = fault.torque_error_front * level if fault.torque_error_front is not None else fault.torque_error_front
            self._sheet.cell(row=self._current_row, column=self._idx_torque_rear_axle).value = fault.torque_error_rear * level if fault.torque_error_rear is not None else fault.torque_error_rear
            self._sheet.cell(row=self._current_row, column=self._idx_torque_slew_rate).value = fault.slew_rate

    def _get_reactions(self, fault, scenario):
        """
Method to get the expected reactions based on the fault and road conditions
        :param fault: Malfunction
        :param scenario: Scenario
        :return: Returns all the expected reactions in a list
        """

        # No reaction is always considered as a separate testrun (for severity calculation):
        reactions = [[VerySlowSteeringReaction(0), SlowSteeringReaction(0), BrakingReaction(0)]]

        # The force of the braking reaction depends on the unintended acceleration level.
        # For higher unintended acceleration a stronger braking reaction is assumed.
        # The braking however is limited to a lower level on icy surfaces.

        if isinstance(fault, TorqueFault):
            if fault.get_overall_torque() > 100:
                braking_reaction = self._config.get_float('Reaction', 'braking_torque_fault_high')
            elif fault.get_overall_torque() < 0:
                braking_reaction = 5
            else:
                braking_reaction = self._config.get_float('Reaction', 'braking_torque_fault_low')
        else:
            braking_reaction = self._config.get_float('Reaction', 'braking_normal')
        if scenario.road_friction <= self._config.get_float('Road_friction', 'icy'):
            braking_reaction = min(braking_reaction, self._config.get_float('Reaction', 'braking_low_friction'))

        # Braking without steering is a reaction that is expected always unless the fault is already leading to a high deceleration
        if not isinstance(fault, TorqueFault) or fault.get_overall_torque() >= 0:
            reactions.append([VerySlowSteeringReaction(0), SlowSteeringReaction(0), BrakingReaction(braking_reaction)])

        # Braking reaction together with steering correction is expected always except when driving on a straight road with a high friction
        friction_limit = self._config.get_float('Road_friction', 'gravel')
        if (not isinstance(scenario.road_radius, str)) or scenario.road_friction < friction_limit or (isinstance(fault, TorqueFault) and fault.losing_stability()):
            # Braking with very slow steering reaction:
            reactions.append([BrakingReaction(braking_reaction), VerySlowSteeringReaction(self._config.get_float('Reaction', 'very_slow_steering'))])
            # Braking with slow steering reaction:
            reactions.append([BrakingReaction(braking_reaction), SlowSteeringReaction(self._config.get_float('Reaction', 'slow_steering'))])
        return reactions

    def save(self):
        """
Formatting the sheet and saving it
        """
        print(f"Status: Saving to {self._path}...")
        for iCol in range(1, self._sheet.max_column):
            font = copy.copy(self._sheet.cell(row=self._header_size + 1, column=iCol).font)
            alignment = copy.copy(self._sheet.cell(row=self._header_size + 1, column=iCol).alignment)
            number_format = self._sheet.cell(row=self._header_size + 1, column=iCol).number_format
            for iRow in range(self._header_size + 2, self._current_row + 1):
                self._sheet.cell(row=iRow, column=iCol).font = font
                self._sheet.cell(row=iRow, column=iCol).alignment = alignment
                self._sheet.cell(row=iRow, column=iCol).number_format = number_format
        if self._mode.lower() == 'ftti_list' or self._mode.lower() == 'acceptance_list':
            self._sheet.column_dimensions['Z'].hidden = True
            self._sheet.column_dimensions['AA'].hidden = False
            self._sheet.column_dimensions['AB'].hidden = False
            self._sheet.column_dimensions['AC'].hidden = True
            self._sheet.column_dimensions['AD'].hidden = True
            self._sheet.column_dimensions['AE'].hidden = True
            self._sheet.column_dimensions['AF'].hidden = True
            self._sheet.column_dimensions['AG'].hidden = True
            self._sheet.column_dimensions['AH'].hidden = True
            self._sheet.column_dimensions['AI'].hidden = False

        self._workbook.save(self._path)
        print('Status: Done')


if __name__ == '__main__':
    print('Status: Started')
    main('Scenario_List')
    # main('FTTI_List')
    # main('Acceptance_List')
